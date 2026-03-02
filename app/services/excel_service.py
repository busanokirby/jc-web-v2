"""
Excel Report Generation Service
Creates .xlsx files with sales and repair data
"""
from __future__ import annotations
from datetime import date
from decimal import Decimal
from typing import Dict, Optional
import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReportService:
    """Service for generating Excel reports"""
    
    # Color scheme
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    SUMMARY_FILL = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    SUMMARY_FONT = Font(bold=True, size=11)
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def generate_filename(start_date: date, end_date: date) -> str:
        """Generate filename based on date range"""
        if start_date == end_date:
            return f"Sales_Report_{start_date.strftime('%Y_%m_%d')}.xlsx"
        else:
            return f"Sales_Report_{start_date.strftime('%Y_%m_%d')}_to_{end_date.strftime('%Y_%m_%d')}.xlsx"
    
    @staticmethod
    def create_report(report_data: Dict) -> Optional[bytes]:
        """
        Create Excel workbook from report data.
        
        CRITICAL: Validates report_data structure before processing.
        Raises exception on validation failure (no silent failures).
        
        Args:
            report_data: Report dict from ReportService.generate_report_data()
        
        Returns:
            Excel file as bytes, or None if validation fails
            
        Raises:
            ValueError: If required fields are missing
            Exception: If workbook creation fails
        """
        # Validate required fields before processing
        required_keys = [
            'date_range', 'frequency', 'total_revenue', 'total_transactions',
            'total_sales_payments', 'total_repair_payments', 'payment_breakdown'
        ]
        missing_keys = [k for k in required_keys if k not in report_data]
        if missing_keys:
            error_msg = f"Report data validation failed: missing keys {missing_keys}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        try:
            wb = Workbook()
            logger.debug("Workbook created")
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
                logger.debug("Default sheet removed")
            
            # Create sheets
            logger.debug("Creating summary sheet...")
            ExcelReportService._create_summary_sheet(wb, report_data)
            logger.debug("Summary sheet created successfully")
            
            logger.debug("Creating transactions sheet...")
            ExcelReportService._create_transactions_sheet(wb, report_data)
            logger.debug("Transactions sheet created successfully")
            
            logger.debug("Creating sales sheet...")
            ExcelReportService._create_sales_sheet(wb, report_data)
            logger.debug("Sales sheet created successfully")
            
            logger.debug("Creating repairs sheet...")
            ExcelReportService._create_repairs_sheet(wb, report_data)
            logger.debug("Repairs sheet created successfully")
            
            # Save to bytes
            logger.debug("Converting workbook to bytes...")
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            excel_bytes = output.getvalue()
            
            if not excel_bytes or len(excel_bytes) == 0:
                raise ValueError("Workbook conversion produced empty bytes")
            
            logger.info(f"Excel report generated successfully: {len(excel_bytes)} bytes")
            return excel_bytes
        
        except Exception as e:
            logger.error(f"Error creating Excel report: {e}", exc_info=True)
            raise  # Re-raise to allow caller to handle failure explicitly
    
    @staticmethod
    def _create_summary_sheet(wb: Workbook, report_data: Dict):
        """Create summary sheet with KPIs"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "SALES & REPAIR REPORT"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Date range
        ws['A3'] = "Report Period:"
        ws['B3'] = report_data.get('date_range', 'N/A')
        ws['A3'].font = Font(bold=True)
        
        ws['A4'] = "Frequency:"
        ws['B4'] = report_data.get('frequency', 'N/A').replace('_', ' ').title()
        ws['A4'].font = Font(bold=True)
        
        # KPIs
        row = 6
        metrics = [
            ("Total Revenue", report_data.get('total_revenue', 0), "₱{:,.2f}"),
            ("Total Transactions", report_data.get('total_transactions', 0), "{}"),
            ("Sales Payments", report_data.get('total_sales_payments', 0), "₱{:,.2f}"),
            ("Repair Payments", report_data.get('total_repair_payments', 0), "₱{:,.2f}"),
        ]
        
        for label, value, fmt in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = fmt.format(value)
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = ExcelReportService.SUMMARY_FILL
            ws[f'B{row}'].fill = ExcelReportService.SUMMARY_FILL
            row += 1
        
        # Payment breakdown
        ws[f'A{row}'] = "Payment Method Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        ws[f'A{row}'] = "Method"
        ws[f'B{row}'] = "Count"
        ws[f'C{row}'] = "Total"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = ExcelReportService.HEADER_FONT
            ws[f'{col}{row}'].fill = ExcelReportService.HEADER_FILL
            ws[f'{col}{row}'].border = ExcelReportService.BORDER
        row += 1
        
        for method, data in sorted(report_data.get('payment_breakdown', {}).items()):
            ws[f'A{row}'] = method
            ws[f'B{row}'] = data.get('count', 0)
            ws[f'C{row}'] = data.get('total', 0)
            ws[f'C{row}'].number_format = '₱#,##0.00'
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
    
    @staticmethod
    def _create_transactions_sheet(wb: Workbook, report_data: Dict):
        """Create combined transactions sheet with all sales and repairs - matching daily_sales.html display"""
        ws = wb.create_sheet("Transactions", 1)
        
        # Header
        headers = ["Customer", "Type", "Description", "Status", "Amount", "Date/Time"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ExcelReportService.HEADER_FONT
            cell.fill = ExcelReportService.HEADER_FILL
            cell.border = ExcelReportService.BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[1].height = 20
        
        # Get transactions - prefer pre-formatted received_records from daily_sales context
        transactions = []
        total_amount = 0
        
        # Use received_records if available (from daily_sales.html context)
        if 'received_records' in report_data and report_data['received_records']:
            received_records = report_data['received_records']
            for rec in received_records:
                # received_records are already formatted with all needed fields
                dt = rec.get('datetime')
                dt_str = ''
                if dt:
                    if hasattr(dt, 'strftime'):
                        dt_str = dt.strftime('%Y-%m-%d %I:%M %p')
                    else:
                        dt_str = str(dt)
                
                transactions.append({
                    'customer': rec.get('customer', ''),
                    'type': rec.get('type', ''),
                    'description': rec.get('description', ''),
                    'status': 'Partial' if rec.get('is_partial') else 'Paid',
                    'amount': rec.get('amount', 0),
                    'datetime_str': dt_str
                })
                total_amount += rec.get('amount', 0)
        else:
            # Fallback: construct from sales_records and repair_records
            # Add sales transactions (filter for received payments only)
            for sale in report_data.get('sales_records', []):
                amount = sale.get('amount_paid', 0) or sale.get('amount', 0)
                if amount > 0:
                    dt_str = ''
                    dt = sale.get('payment_date')
                    if dt:
                        if hasattr(dt, 'strftime'):
                            dt_str = dt.strftime('%Y-%m-%d %I:%M %p')
                        else:
                            dt_str = str(dt)
                    
                    transactions.append({
                        'customer': sale.get('customer_name', ''),
                        'type': 'Sale',
                        'description': sale.get('items_description', ''),
                        'status': 'Partial' if sale.get('payment_status', '').upper() == 'PARTIAL' else 'Paid',
                        'amount': amount,
                        'datetime_str': dt_str
                    })
                    total_amount += amount
            
            # Add repair transactions (filter for received payments only)
            for repair in report_data.get('repair_records', []):
                amount = repair.get('amount_paid', 0) or repair.get('amount', 0)
                if amount > 0:
                    dt_str = ''
                    dt = repair.get('payment_date')
                    if dt:
                        if hasattr(dt, 'strftime'):
                            dt_str = dt.strftime('%Y-%m-%d %I:%M %p')
                        else:
                            dt_str = str(dt)
                    
                    transactions.append({
                        'customer': repair.get('customer_name', ''),
                        'type': 'Repair',
                        'description': repair.get('device_type', ''),
                        'status': 'Partial' if repair.get('payment_status', '').upper() == 'PARTIAL' else 'Paid',
                        'amount': amount,
                        'datetime_str': dt_str
                    })
                    total_amount += amount
        
        # Write data
        row = 2
        for trans in transactions:
            ws.cell(row=row, column=1, value=trans['customer'])  # Customer
            ws.cell(row=row, column=2, value=trans['type'])  # Type
            ws.cell(row=row, column=3, value=trans['description'])  # Description
            ws.cell(row=row, column=4, value=trans['status'])  # Status
            
            amount_cell = ws.cell(row=row, column=5, value=trans['amount'])
            amount_cell.number_format = '₱#,##0.00'
            
            ws.cell(row=row, column=6, value=trans['datetime_str'])  # Date/Time
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = ExcelReportService.BORDER
            
            row += 1
        
        # Grand total row
        if row > 2:
            row += 1  # Add blank row for spacing
            
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).fill = ExcelReportService.SUMMARY_FILL
            
            total_cell = ws.cell(row=row, column=5, value=total_amount)
            total_cell.number_format = '₱#,##0.00'
            total_cell.font = Font(bold=True, size=11)
            total_cell.fill = ExcelReportService.SUMMARY_FILL
            
            # Style the entire total row
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = ExcelReportService.SUMMARY_FILL
                ws.cell(row=row, column=col).border = ExcelReportService.BORDER
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
    
    @staticmethod
    def _create_sales_sheet(wb: Workbook, report_data: Dict):
        """Create sales transactions sheet"""
        ws = wb.create_sheet("Sales", 2)
        
        # Header
        headers = ["Invoice #", "Customer", "Payment Method", "Amount Paid", "Payment Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ExcelReportService.HEADER_FONT
            cell.fill = ExcelReportService.HEADER_FILL
            cell.border = ExcelReportService.BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[1].height = 20
        
        # Data
        row = 2
        for sale in report_data.get('sales_records', []):
            ws.cell(row=row, column=1, value=sale.get('invoice_number', ''))
            ws.cell(row=row, column=2, value=sale.get('customer_name', ''))
            ws.cell(row=row, column=3, value=sale.get('payment_method', ''))
            
            amount_cell = ws.cell(row=row, column=4, value=sale.get('amount_paid', 0))
            amount_cell.number_format = '₱#,##0.00'
            
            date_cell = ws.cell(row=row, column=5, value=sale.get('payment_date', ''))
            date_cell.number_format = 'yyyy-mm-dd'
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = ExcelReportService.BORDER
            
            row += 1
        
        # Totals row
        if row > 2:
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            total_cell = ws.cell(row=row, column=4, value=report_data.get('total_sales_payments', 0))
            total_cell.number_format = '₱#,##0.00'
            total_cell.font = Font(bold=True)
            total_cell.fill = ExcelReportService.SUMMARY_FILL
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    @staticmethod
    def _create_repairs_sheet(wb: Workbook, report_data: Dict):
        """Create repairs transactions sheet"""
        ws = wb.create_sheet("Repairs", 3)
        
        # Header
        headers = ["Ticket #", "Customer", "Device", "Payment Method", "Amount Paid", "Payment Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ExcelReportService.HEADER_FONT
            cell.fill = ExcelReportService.HEADER_FILL
            cell.border = ExcelReportService.BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[1].height = 20
        
        # Data
        row = 2
        for repair in report_data.get('repair_records', []):
            ws.cell(row=row, column=1, value=repair.get('ticket_number', ''))
            ws.cell(row=row, column=2, value=repair.get('customer_name', ''))
            ws.cell(row=row, column=3, value=repair.get('device_type', ''))
            ws.cell(row=row, column=4, value=repair.get('payment_method', ''))
            
            amount_cell = ws.cell(row=row, column=5, value=repair.get('amount_paid', 0))
            amount_cell.number_format = '₱#,##0.00'
            
            date_cell = ws.cell(row=row, column=6, value=repair.get('payment_date', ''))
            date_cell.number_format = 'yyyy-mm-dd'
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = ExcelReportService.BORDER
            
            row += 1
        
        # Totals row
        if row > 2:
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            total_cell = ws.cell(row=row, column=5, value=report_data.get('total_repair_payments', 0))
            total_cell.number_format = '₱#,##0.00'
            total_cell.font = Font(bold=True)
            total_cell.fill = ExcelReportService.SUMMARY_FILL
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
